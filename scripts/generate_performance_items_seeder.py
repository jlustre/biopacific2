import json
import re
from pathlib import Path

import openpyxl

BASE = Path(r"c:\bio-pacific\website\ COMPETENCY_SKILLS CHECKS - ALL DEPARTMENTS\Annual Performance Evaluation")
OUT_JSON = Path(__file__).resolve().parents[1] / "storage/app/performance_items_merged.json"
OUT_SEEDER = Path(__file__).resolve().parents[1] / "database/seeders/EmployeePerformanceItemsSeeder.php"

FILES = {
    "rn_lvn": "Performance Appraisal - RN & LVN.xlsx",
    "cna": "Performance Appraisal - CNA.xlsx",
    "dietary_aide": "Performance Appraisal - Dietary Aide.xlsx",
    "general_services": "Performance Appraisal - General Services(Non-Management).xlsx",
    "housekeeper": "Performance Appraisal - Housekeeper.xlsx",
    "laundry_aide": "Performance Appraisal - Laundry Aide.xlsx",
    "maintenance": "Performance Appraisal - Maintenance.xlsx",
    "management": "Performance Appraisal - Management.xlsx",
}

ROMAN_SECTION = re.compile(r"^[IVXLC]+\.\s+(.+)$", re.I)
STOP_SECTION = re.compile(
    r"^(areas for development|areas requiring|development plans|employee comments|total of all sections|overall performance|employee:|reviewer:|supervisor:|the following signatures)",
    re.I,
)
SKIP_ITEM = re.compile(
    r"^(comments:?|section [ivx]+ rating|section [ivx]+ total|total|rating description|codes|employee:|reviewer:|supervisor:)$",
    re.I,
)

SECTION_MAP = {
    "JOB SKILLS AND KNOWLEDGE": "Job Skills and Knowledge",
    "DEPENDABILITY": "Dependability",
    "INTERPERSONAL SKILLS": "Interpersonal Skills",
    "ORGANIZATIONAL SKILLS": "Organizational Skills",
    "COMMUNICATION SKILLS": "Communication Skills",
    "PROBLEM SOLVING": "Problem-Solving Skills",
    "PROBLEM-SOLVING SKILLS": "Problem-Solving Skills",
    "SAFETY & HEALTH": "Safety & Health",
    "KNOWLEDGE & EXPERIENCE": "Knowledge & Experience",
    "ORGANIZATION SKILLS": "Organization Skills",
    "DECISION MAKING SKILLS": "Decision Making Skills",
    "LEADERSHIP SKILLS": "Leadership Skills",
    "PERSONAL INTEGRITY": "Personal Integrity",
}

# Every position title in the app maps to exactly one appraisal template (Excel file).
POSITION_TITLE_TO_TEMPLATE = {
    "Registered Nurse": "rn_lvn",
    "Licensed Vocational Nurse": "rn_lvn",
    "Licensed Nurse": "rn_lvn",
    "IP Nurse": "rn_lvn",
    "Certified Nursing Assistant": "cna",
    "Nursing Assistant": "cna",
    "Dietary Aide": "dietary_aide",
    "Cook": "dietary_aide",
    "Housekeeper": "housekeeper",
    "Janitor": "housekeeper",
    "Laundry Staff": "laundry_aide",
    "Maintenance Technician": "maintenance",
    "Administrator": "management",
    "Activities Director": "management",
    "Business Office Manager": "management",
    "Medical Records Director": "management",
    "Food Services Director": "management",
    "Dietary Manager": "management",
    "Housekeeping Supervisor": "management",
    "Maintenance Director": "management",
    "Marketing Director": "management",
    "Social Services Director": "management",
    "Rehab Manager": "management",
    "Director of Staff Development": "management",
    "Director of Nursing": "management",
    "Charge Nurse": "management",
    "Activity Assistant": "general_services",
    "Admissions Coordinator": "general_services",
    "Receptionist": "general_services",
    "Office Staff": "general_services",
    "Other": "general_services",
    "Medical Records Clerk": "general_services",
    "Social Worker": "general_services",
    "Resident Liaison": "general_services",
    "Case Manager": "general_services",
    "MDS Coordinator": "general_services",
    "Occupational Therapist": "general_services",
    "OT/PT Assistant": "general_services",
    "Physical Therapist": "general_services",
    "Unit Clerk": "general_services",
    "Staff Development Coordinator": "general_services",
}


def norm_section(raw: str) -> str | None:
    text = re.sub(r"\s+", " ", raw.strip())
    match = ROMAN_SECTION.match(text)
    label = match.group(1).strip().upper() if match else text.upper()
    return SECTION_MAP.get(label)


def clean_item(raw) -> str | None:
    if not raw or not isinstance(raw, str):
        return None
    text = re.sub(r"\s+", " ", raw.strip())
    if len(text) < 8:
        return None
    if SKIP_ITEM.search(text):
        return None
    if STOP_SECTION.search(text):
        return None
    if text.startswith(("PURPOSE", "WHEN TO", "HOW TO", "The employee exceeds", "The employee meets", "The employee has failed")):
        return None
    if "PERFORMANCE AREAS" in text:
        return None
    if text.startswith("Use the number from the shaded box"):
        return None
    return text


def extract_template_items(category: str) -> list[tuple[str, str]]:
    ws = openpyxl.load_workbook(BASE / FILES[category], data_only=True).active
    section = None
    in_performance_areas = False
    items: list[tuple[str, str]] = []

    for row in range(1, ws.max_row + 1):
        cell = ws.cell(row, 1).value
        if not cell or not isinstance(cell, str):
            continue
        text = re.sub(r"\s+", " ", cell.strip())

        if "PERFORMANCE AREAS" in text.upper():
            in_performance_areas = True
            continue

        if not in_performance_areas:
            continue

        if STOP_SECTION.search(text):
            break

        normalized_section = norm_section(text)
        if normalized_section:
            section = normalized_section
            continue

        if not section:
            continue

        item = clean_item(text)
        if item:
            items.append((section, item))

    return items


def build_merged_items() -> list[dict]:
    merged: dict[tuple[str, str], set[str]] = {}
    section_order: list[str] = []

    for category in FILES:
        for section, item in extract_template_items(category):
            key = (section, item)
            merged.setdefault(key, set()).add(category)
            if section not in section_order:
                section_order.append(section)

    # Ensure management-only sections appear after standard sections.
    for section in [
        "Knowledge & Experience",
        "Organization Skills",
        "Decision Making Skills",
        "Leadership Skills",
        "Personal Integrity",
    ]:
        if section not in section_order:
            section_order.append(section)

    rows = []
    for section in section_order:
        section_items = sorted(
            [(item, cats) for (sec, item), cats in merged.items() if sec == section],
            key=lambda pair: pair[0].lower(),
        )
        for item, categories in section_items:
            rows.append(
                {
                    "section": section,
                    "item": item,
                    "categories": sorted(categories),
                }
            )

    return rows


def php_str(value: str) -> str:
    return value.replace("\\", "\\\\").replace("'", "\\'")


def template_titles_php() -> str:
    by_template: dict[str, list[str]] = {key: [] for key in FILES}
    for title, template in POSITION_TITLE_TO_TEMPLATE.items():
        by_template.setdefault(template, []).append(title)

    lines = ["        return ["]
    for template, titles in by_template.items():
        quoted = ", ".join(f"'{php_str(title)}'" for title in sorted(titles))
        lines.append(f"            '{template}' => [{quoted}],")
    lines.append("        ];")
    return "\n".join(lines)


def title_map_php() -> str:
    lines = ["        return ["]
    for title, template in sorted(POSITION_TITLE_TO_TEMPLATE.items()):
        lines.append(f"            '{php_str(title)}' => '{template}',")
    lines.append("        ];")
    return "\n".join(lines)


def write_seeder(items: list[dict]) -> None:
    lines = [
        "<?php",
        "",
        "namespace Database\\Seeders;",
        "",
        "use App\\Models\\EmployeePerformanceItem;",
        "use App\\Support\\PerformanceAppraisalTemplate;",
        "use Illuminate\\Database\\Seeder;",
        "use Illuminate\\Support\\Facades\\DB;",
        "",
        "class EmployeePerformanceItemsSeeder extends Seeder",
        "{",
        "    protected function getPositionIdsByTitles(array|string $titles): array",
        "    {",
        "        $titles = array_values(array_filter(array_map('trim', (array) $titles)));",
        "",
        "        if ($titles === []) {",
        "            return [];",
        "        }",
        "",
        "        return DB::table('positions')",
        "            ->whereIn('title', $titles)",
        "            ->orderBy('id')",
        "            ->pluck('id')",
        "            ->map(fn ($id) => (int) $id)",
        "            ->all();",
        "    }",
        "",
        "    protected function seedPerformanceItems(array $items, int &$order): void",
        "    {",
        "        foreach ($items as $row) {",
        "            $positionIds = [];",
        "            foreach ($row['categories'] as $category) {",
        "                $positionIds = array_merge(",
        "                    $positionIds,",
        "                    PerformanceAppraisalTemplate::positionIdsForTemplate($category)",
        "                );",
        "            }",
        "",
        "            $positionIds = array_values(array_unique(array_map('intval', $positionIds)));",
        "",
        "            EmployeePerformanceItem::query()->updateOrCreate(",
        "                [",
        "                    'section' => $row['section'],",
        "                    'item' => $row['item'],",
        "                ],",
        "                [",
        "                    'position_ids' => $positionIds,",
        "                    'order' => $order,",
        "                ]",
        "            );",
        "",
        "            $order++;",
        "        }",
        "    }",
        "",
        "    public function run(): void",
        "    {",
        "        EmployeePerformanceItem::query()->delete();",
        "",
        "        $order = 0;",
        "        $items = [",
    ]

    for row in items:
        cats = ", ".join(f"'{c}'" for c in row["categories"])
        lines.extend(
            [
                "            [",
                f"                'section' => '{php_str(row['section'])}',",
                f"                'item' => '{php_str(row['item'])}',",
                f"                'categories' => [{cats}],",
                "            ],",
            ]
        )

    lines.extend(
        [
            "        ];",
            "",
            "        $this->seedPerformanceItems($items, $order);",
            "    }",
            "}",
            "",
        ]
    )

    OUT_SEEDER.write_text("\n".join(lines), encoding="utf-8", newline="\n")


def write_template_class() -> None:
    path = Path(__file__).resolve().parents[1] / "app/Support/PerformanceAppraisalTemplate.php"
    titles_block = template_titles_php().replace("        return [", "        return [", 1)
    map_block = title_map_php()

    content = f"""<?php

namespace App\\Support;

use App\\Models\\Position;
use Illuminate\\Support\\Facades\\DB;

class PerformanceAppraisalTemplate
{{
    /**
     * Appraisal template keys sourced from the Excel workbooks.
     */
    public const RN_LVN = 'rn_lvn';
    public const CNA = 'cna';
    public const DIETARY_AIDE = 'dietary_aide';
    public const GENERAL_SERVICES = 'general_services';
    public const HOUSEKEEPER = 'housekeeper';
    public const LAUNDRY_AIDE = 'laundry_aide';
    public const MAINTENANCE = 'maintenance';
    public const MANAGEMENT = 'management';

    /**
     * @return array<string, string>
     */
    public static function positionTitleToTemplate(): array
    {{
{map_block}
    }}

    /**
     * @return array<string, list<string>>
     */
    public static function templatePositionTitles(): array
    {{
{titles_block}
    }}

    public static function templateForPositionTitle(?string $positionTitle): ?string
    {{
        if (! filled($positionTitle)) {{
            return null;
        }}

        return static::positionTitleToTemplate()[trim($positionTitle)] ?? null;
    }}

    /**
     * @return list<int>
     */
    public static function positionIdsForTemplate(string $template): array
    {{
        $titles = static::templatePositionTitles()[$template] ?? [];

        if ($titles === []) {{
            return [];
        }}

        return DB::table('positions')
            ->whereIn('title', $titles)
            ->orderBy('id')
            ->pluck('id')
            ->map(fn ($id) => (int) $id)
            ->all();
    }}

    /**
     * @return list<int>
     */
    public static function positionIdsForPositionTitle(?string $positionTitle): array
    {{
        $template = static::templateForPositionTitle($positionTitle);

        return $template ? static::positionIdsForTemplate($template) : [];
    }}

    public static function positionIdForTitle(?string $positionTitle): ?int
    {{
        if (! filled($positionTitle)) {{
            return null;
        }}

        $id = Position::query()->where('title', trim($positionTitle))->value('id');

        return $id ? (int) $id : null;
    }}
}}
"""
    path.write_text(content, encoding="utf-8", newline="\n")


def main() -> None:
    items = build_merged_items()
    OUT_JSON.write_text(
        json.dumps({"items": items, "position_title_to_template": POSITION_TITLE_TO_TEMPLATE}, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    write_template_class()
    write_seeder(items)

    print(f"Merged unique items: {len(items)}")
    for category in FILES:
        count = sum(1 for row in items if category in row["categories"])
        print(f"  {category}: {count} items tagged")

    print(f"Wrote {OUT_SEEDER}")


if __name__ == "__main__":
    main()
